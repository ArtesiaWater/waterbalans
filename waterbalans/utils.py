"""This file contains practical classes and methods for use throughout the
"Waterbalans" model."""
import os

import numpy as np
import pandas as pd
from pandas import to_datetime, to_timedelta


def excel2datetime(excel_datenum, freq="D", start_date="1899-12-30"):
    """Method to convert excel datetime to pandas timetime objects.

    Parameters
    ----------
    excel_datenum: datetime index
        can be a datetime object or a pandas datetime index.
    freq:

    Returns
    -------
    datetimes: pandas.datetimeindex
    """
    datetimes = to_datetime(start_date) + to_timedelta(excel_datenum, freq)
    return datetimes


def makkink_to_penman(e, use_excel_factors=False):
    """Method to transform the the makkink potential evaporation to Penman
    evaporation for open water.

    Parameters
    ----------
    e: pandas.Series
        Pandas Series containing the evaporation with the date as index.
    use_excel_factors: bool, optional default is False
        if True, uses the excel factors, only difference is that evaporation
        in december is 0.0.

    Returns
    -------
    e: pandas.Series
        Penman evaporation as a pandas time series object.

    Notes
    -----
    Van Penman naar Makkink, een nieuwe berekeningswijze voor de
    klimatologische verdampingsgetallen, KNMI/CHO, rapporten en nota's, no.19
    """
    if use_excel_factors:
        # penman = [2.500, 1.071, 0.789, 0.769, 0.769, 0.763, 0.789, 0.838, 0.855,
        #           1.111, 1.429, np.inf]  # col E47:E59 in Excel e_r / e_o, with 0 evap in december.
        penman = 1.0 / np.array(
            [
                0.4,
                0.933333333,
                1.266666667,
                1.3,
                1.3,
                1.310000000,
                1.266666667,
                1.193333333,
                1.170000000,
                0.9,
                0.7,
                0.0,
            ]
        )
    else:
        penman = [
            2.500,
            1.071,
            0.789,
            0.769,
            0.769,
            0.763,
            0.789,
            0.838,
            0.855,
            1.111,
            1.429,
            1.000,
        ]  # col E47:E59 in Excel e_r / e_o
    e = e.copy()
    for i in range(1, 13):
        with np.errstate(divide="ignore"):
            e.loc[e.index.month == i] = (
                e.loc[e.index.month == i] / penman[i - 1]
            )  # for first list
    return e


def calculate_cso(prec, Bmax, POCmax, alphasmooth=0.1):
    """Calculate Combined Sewer Overflow timeseries based on hourly
    precipitation series.

    Parameters
    ----------
    prec : pd.Series
        hourly precipitation
    Bmax : float
        maximum storage capacity in meters
    POCmax : float
        maximum pump (over) capacity
    alphasmooth : float, optional
        factor for exponential smoothing (the default is 0.1)

    Returns
    -------
    pd.Series
        timeseries of combined sewer overflows (cso)
    """

    p_smooth = prec.ewm(alpha=alphasmooth, adjust=False).mean()
    b = p_smooth.copy()
    poc = p_smooth.copy()
    cso = p_smooth.copy()
    vol = p_smooth.copy()

    for i in range(1, len(p_smooth.index)):
        vol.iloc[i] = p_smooth.iloc[i] + b.iloc[i - 1] - poc.iloc[i - 1]
        b.iloc[i] = np.min([vol.iloc[i], Bmax])
        poc.iloc[i] = np.min([b.iloc[i], POCmax])
        cso.iloc[i] = np.max([vol.iloc[i] - Bmax, 0.0])

    cso_daily = cso.resample("D").sum()

    return cso_daily


def get_model_input_from_excel(excelfile):
    """get modelstructure, timeseries, and parameters from an excel file. The
    structure of the excelfile is defined. See example file at
    https://github.com/ArtesiaWater/waterbalans/tree/master/voorbeelden/data.

    Parameters
    ----------
    excelfile : str, path to excel file
        Path to excelfile

    Returns
    -------
    df_ms, df_ts, df_params: pandas.DataFrames
        DataFrames containing info about modelstructure,
        timeseries and parameters.
    """
    xls = pd.ExcelFile(excelfile, engine="openpyxl")

    df_ms = pd.read_excel(
        xls, sheet_name="modelstructure", skiprows=[1], header=[0], index_col=None
    )
    df_ts = pd.read_excel(
        xls,
        sheet_name="reeksen",
        skiprows=[1],
        header=[0],
        index_col=None,
        usecols="A:J",
    )
    df_params = pd.read_excel(
        xls,
        sheet_name="parameters",
        skiprows=[1],
        header=[0],
        index_col=None,
        usecols="A:G",
    )

    return df_ms, df_ts, df_params


def get_extra_series_from_excel(excelfile, sheet_name="extra_reeksen"):
    """Load extra timeseries from excelfile containing all info for an EAG. The
    structure of the excelfile is defined. See example file at
    https://github.com/ArtesiaWater/waterbalans/tree/master/voorbeelden/data.

    Parameters
    ----------
    excelfile : str, path to excelfile
        path to excelfile

    Returns
    -------
    df_series: pandas.DataFrame
        DataFrame containing series to be added to waterbalance
    """

    xls = pd.ExcelFile(excelfile, engine="openpyxl")
    df_series = pd.read_excel(
        xls,
        sheet_name=sheet_name,
        skiprows=[1],
        header=[0],
        index_col=[0],
        parse_dates=True,
    )

    return df_series


def get_wqparams_from_excel(excelfile, sheet_name="stoffen"):
    """Load water quality parameters from excelfile containing all info for an
    EAG. The structure of the excelfile is defined. See example file at
    https://github.com/ArtesiaWater/waterbalans/tree/master/voorbeelden/data.

    Parameters
    ----------
    excelfile : str, path to excelfile
        path to excelfile

    Returns
    -------
    df_series: pandas.DataFrame
        DataFrame containing water quality parameters
    """

    xls = pd.ExcelFile(excelfile, engine="openpyxl")
    df_series = pd.read_excel(
        xls,
        sheet_name=sheet_name,
        skiprows=[1],
        header=[0],
        index_col=None,
        usecols="A:I",
    )

    return df_series


def get_extra_series_from_pickle(picklefile, compression="zip"):
    """Load timeseries from pickle.

    Parameters
    ----------
    picklefile : str, path to picklefile
        path to picklefile

    Returns
    -------
    df_series: pandas.DataFrame
        DataFrame containing series
    """

    df_series = pd.read_pickle(picklefile, compression=compression)
    return df_series


def add_timeseries_to_obj(
    eag_or_gaf, df, tmin=None, tmax=None, overwrite=False, data_from_excel=False
):
    """Add timeseries to EAG or GAF. Only parses column names starting with
    'Neerslag', 'Verdamping', 'Inlaat', 'Uitlaat', 'Peil', or 'Gemaal'.

    Parameters
    ----------
    eag_or_gaf : waterbalans.Eag or waterbalans.Gaf
        An existing Eag or Gaf object
    df : pandas.DataFrame
        DataFrame with DateTimeIndex containing series to be added
    tmin : pandas.TimeStamp, optional
        start time for added series (the default is None, which
        attempts to pick up tmin from existing Eag or Gaf object)
    tmax : pandas.TimeStamp, optional
        end time for added series (the default is None, which
        attempts to pick up tmax from existing Eag or Gaf object)
    overwrite : bool, optional
        overwrite series if name already exists in Eag or Gaf object (the default is False)
    data_from_excel: bool, optional
        if True, assumes data source is Excel Balance 'uitgangspunten' sheet.
        Function will make an assumption about the column names and order and
        disregards column names of the passed DataFrame. If False uses DataFrame
        column names (default).
    """
    o = eag_or_gaf
    # if not isinstance(o, Eag) or not isinstance(o, Gaf):
    #     raise ValueError(
    #         "Arg 'eag_or_gaf' must be Eag or Gaf object. Received {}".format(type(o)))

    try:
        if tmin is None:
            tmin = o.series.index[0]
        if tmax is None:
            tmax = o.series.index[-1]
    except IndexError:
        raise ValueError("tmin/tmax cannot be inferred from EAG/GAF object.")

    if data_from_excel:
        columns = [
            "neerslag",
            "verdamping",
            "peil",
            "Gemaal1",
            "Gemaal2",
            "Gemaal3",
            "Gemaal4",
            "Inlaat voor calibratie",
            "gemengd gerioleerd stelsel",
            "Inlaat1",
            "Inlaat2",
            "Inlaat3",
            "Inlaat4",
            "Uitlaat1",
            "Uitlaat2",
            "Uitlaat3",
            "Uitlaat4",
        ]
    else:
        columns = df.columns

    eag_series = o.series.columns

    # Inlaat/Uitlaat
    factor = 1.0
    for inam in ["Gemaal", "Inlaat", "Uitlaat"]:
        colmask = [
            True if icol.lower().startswith(inam.lower()) else False for icol in columns
        ]
        series = df.loc[:, colmask]
        # Water bucket converts outgoing fluxes to negative, so outgoing fluxes can be entered positive
        for jcol in range(series.shape[1]):
            # Check if empty
            if series.iloc[:, jcol].dropna().empty:
                o.logger.warning(
                    "'{}' is empty. Continuing...".format(series.columns[jcol])
                )
                continue
            # Check if series is already in EAG
            if data_from_excel:
                colnam = np.array(columns)[colmask][jcol]
            else:
                colnam = series.columns[jcol].split("|")[0]
            if colnam in eag_series:
                if overwrite:
                    o.add_timeseries(
                        factor * series.iloc[:, jcol],
                        name="{}".format(colnam),
                        tmin=tmin,
                        tmax=tmax,
                        fillna=True,
                        method=0.0,
                    )
                else:
                    o.logger.warning(
                        "'{}' already in EAG. No action taken.".format(colnam)
                    )
            else:
                # o.logger.info("Adding '{}' series to EAG.".format(
                #     colnam))
                o.add_timeseries(
                    factor * series.iloc[:, jcol],
                    name="{}".format(colnam),
                    tmin=tmin,
                    tmax=tmax,
                    fillna=True,
                    method=0.0,
                )

    # Peil
    colmask = [True if icol.lower().startswith("peil") else False for icol in columns]
    if np.sum(colmask) > 0:
        peil = df.loc[:, colmask]
        if "Peil" in eag_series:
            if overwrite:
                o.add_timeseries(
                    peil, name="Peil", tmin=tmin, tmax=tmax, fillna=True, method="ffill"
                )
            else:
                o.logger.warning("'Peil' already in EAG. No action taken.")
        else:
            # o.logger.info("Adding 'Peil' series to EAG.")
            o.add_timeseries(
                peil, name="Peil", tmin=tmin, tmax=tmax, fillna=True, method="ffill"
            )

    # q_cso MengRiool overstortreeks
    colmask = [True if icol.lower().startswith("q_cso") else False for icol in columns]
    if np.sum(colmask) > 0:
        q_cso = df.loc[:, colmask] / 100**2
        if "q_cso" in eag_series:
            if overwrite:
                o.add_timeseries(
                    q_cso, name="q_cso", tmin=tmin, tmax=tmax, fillna=True, method=0.0
                )
            else:
                o.logger.warning("'q_cso' already in EAG. No action taken.")
        else:
            # o.logger.info("Adding 'q_cso' series to EAG.")
            o.add_timeseries(
                q_cso, name="q_cso", tmin=tmin, tmax=tmax, fillna=True, method=0.0
            )

    # Neerslag/Verdamping
    for inam in ["Neerslag", "Verdamping"]:
        colmask = [
            True if icol.lower().startswith(inam.lower()) else False for icol in columns
        ]
        if np.sum(colmask) > 0:
            pe = df.loc[:, colmask] * 1e-3
            if inam in eag_series:
                if overwrite:
                    o.add_timeseries(
                        pe, name=inam, tmin=tmin, tmax=tmax, fillna=True, method=0.0
                    )
                else:
                    o.logger.warning(
                        "'{}' already in EAG. No action taken.".format(inam)
                    )
            else:
                # o.logger.info("Adding '{}' series to EAG.".format(inam))
                o.add_timeseries(
                    pe, name=inam, tmin=tmin, tmax=tmax, fillna=True, method=0.0
                )


def create_csvfile_table(csvdir):
    """Creates a DataFrame containing all csv file names for EAGs or GAFS in
    that folder.

    Parameters
    ----------
    csvdir : path to dir
        folder in which csvs are stored

    Returns
    -------
    pandas.DataFrame
        DataFrame containing each CSV for a specific EAG or GAF
    """

    files = [i for i in os.listdir(csvdir) if i.endswith(".csv")]
    eag_df = pd.DataFrame(data=files, columns=["filenames"])
    eag_df["ID"] = eag_df.filenames.apply(
        lambda s: s.split("_")[2].split(".")[0]
        if not s.startswith("stoffen")
        else s.split("_")[3].split(".")[0]
    )
    eag_df["type"] = eag_df.filenames.apply(
        lambda s: s.split("_")[0]
        if not s.startswith("stoffen")
        else "_".join(s.split("_")[:2])
    )
    eag_df.drop_duplicates(subset=["ID", "type"], keep="last", inplace=True)
    file_df = eag_df.pivot(index="ID", columns="type", values="filenames")
    file_df.dropna(how="any", subset=["opp", "param", "reeks"], axis=0, inplace=True)
    return file_df


def compare_to_excel_balance(e, pickle_dir, **kwargs):
    # Read Excel Balance Data (see scrape_excelbalansen.py for details)
    excelbalance = pd.read_pickle(
        os.path.join(pickle_dir, "{}_wbalance.pklz".format(e.name)), compression="zip"
    )
    for icol in excelbalance.columns:
        excelbalance.loc[:, icol] = pd.to_numeric(excelbalance[icol], errors="coerce")

    # Waterbalance comparison
    fig = e.plot.compare_fluxes_to_excel_balance(excelbalance, **kwargs)

    return fig


def eag_params_to_excel_dict(eag):

    return_dicts = []

    index = eag.water.hTargetSeries.index
    # find first (almost) full year
    m0 = index[0].month
    if m0 > 1:
        y0 = index[0].year + 1
    else:
        y0 = index[0].year

    # waterpeilen
    peilen = {
        # min peil 1
        "C64": eag.water.hTargetSeries.loc["31-03-{}".format(y0), "hTargetMin"],
        # min peil 2
        "C65": eag.water.hTargetSeries.loc["15-05-{}".format(y0), "hTargetMin"],
        # min peil 3
        "C66": eag.water.hTargetSeries.loc["30-08-{}".format(y0), "hTargetMin"],
        # min peil 4
        "C67": eag.water.hTargetSeries.loc["15-10-{}".format(y0), "hTargetMin"],
        # max peil 1
        "E64": eag.water.hTargetSeries.loc["31-03-{}".format(y0), "hTargetMax"],
        # max peil 2
        "E65": eag.water.hTargetSeries.loc["15-05-{}".format(y0), "hTargetMax"],
        # max peil 3
        "E66": eag.water.hTargetSeries.loc["30-08-{}".format(y0), "hTargetMax"],
        # max peil 4
        "E67": eag.water.hTargetSeries.loc["15-10-{}".format(y0), "hTargetMax"],
        # min peil altijd
        "C68": eag.parameters.at["hTargetMin_1", "Waarde"],
        # polder peil / start peil
        "D68": eag.parameters.at["hTarget_1", "Waarde"],
        # max peil altijd
        "E68": eag.parameters.at["hTargetMax_1", "Waarde"],
    }
    return_dicts.append(peilen)

    # bakjes
    bp = eag.get_bucket_params()
    buckets = bp.columns.get_level_values(0)

    n_verhard = 0
    n_drain = 0
    n_onverhard = 0
    for b in buckets:
        if b.startswith("Verhard"):
            if n_verhard < 1:
                bv = bp.loc[:, b].squeeze()
                sv = eag.buckets[int(b.split("_")[-1])].series
                verhard = {
                    "F37": bv["RFacOut_2"],
                    "F38": bv["RFacIn_2"],
                    "F39": bv["por_2"],
                    "F40": bv["hMax_1"],
                    # Qkwel_zomer
                    "B39": sv.loc["01-07-{}".format(y0), "Qkwel"],
                    # Qkwel_winter
                    "C39": sv.loc["01-12-{}".format(y0), "Qkwel"],
                    "F43": bv["hInit_1"],
                    "Rekenblad|B5": bv["EFacMax_1"],
                    "Rekenblad|B6": bv["EFacMin_1"],
                    # "Rekenblad|C5": bv["EFacMax_2"],
                    # "Rekenblad|C6": bv["EFacMin_2"],
                }
                return_dicts.append(verhard)
                n_verhard += 1
            else:
                print("Warning: Only 1 Verhard bucket can be written to Excel!")
        if b.startswith("Drain"):
            if n_drain < 1:
                bd = bp.loc[:, b].squeeze()
                sd = eag.buckets[int(b.split("_")[-1])].series
                gedraineerd = {
                    "G37": bd["RFacOut_1"],  # RFacOut laag boven (1)
                    "H37": bd["RFacOut_2"],  # RFacOut laag onder (2)
                    "G38": bd["RFacIn_1"],  # RFacIn laag boven (1)
                    "H38": bd["RFacIn_2"],  # RFacIn laag onder (2)
                    "G39": bd["por_1"],  # por laag boven (1)
                    "H39": bd["por_2"],  # por laag onder (2)
                    "G40": bd["hMax_1"],  # hMax laag boven (1)
                    "H40": bd["hMax_2"],  # hMax laag onder (2)
                    # Qkwel_zomer
                    "B40": sd.loc["01-07-{}".format(y0), "Qkwel"],
                    # Qkwel_winter,
                    "C40": sd.loc["01-12-{}".format(y0), "Qkwel"],
                    "G43": bd["hInit_1"],
                    "H43": bd["hInit_2"],
                    "Rekenblad|D5": bd["EFacMax_1"],  # EFacMax_boven
                    # "Rekenblad|E5": bd["EFacMax_2"],  # EFacMax_onder
                    "Rekenblad|D6": bd["EFacMin_1"],  # EFacMin_boven
                    # "Rekenblad|E6": bd["EFacMin_2"],  # EFacMin_onder
                }
                return_dicts.append(gedraineerd)
                n_drain += 1
            else:
                print("Warning: Only 1 Drain bucket can be written to Excel!")
        if b.startswith("Onverhard"):
            bo = bp.loc[:, b].squeeze()
            so = eag.buckets[int(b.split("_")[-1])].series
            if n_onverhard == 0:
                onverhard1 = {
                    "I37": bo["RFacOut_1"],  # RFacOut bakje 1
                    "I38": bo["RFacIn_1"],  # RFacIn bakje 1
                    "I39": bo["por_1"],  # por bakje 1
                    "I40": bo["hMax_1"],  # hMax bakje 1
                    # Qkwel zomer bakje 1
                    "B41": so.loc["01-07-{}".format(y0), "Qkwel"],
                    # Qkwel winter bakje 1
                    "C41": so.loc["01-12-{}".format(y0), "Qkwel"],
                    "I43": bo["hInit_1"],
                    "Rekenblad|F5": bo["EFacMax_1"],
                    "Rekenblad|F6": bo["EFacMin_1"],
                }
                return_dicts.append(onverhard1)
            elif n_onverhard == 1:
                onverhard2 = {
                    "J37": bo["RFacOut_1"],  # RFacOut bakje 2
                    "J38": bo["RFacIn_1"],  # RFacIn bakje 2
                    "J39": bo["por_1"],  # por bakje 2
                    "J40": bo["hMax_1"],  # hMax bakje 2
                    # Qkwel zomer bakje 2
                    "B42": so.loc["01-07-{}".format(y0), "Qkwel"],
                    # Qkwel winter bakje 2
                    "C42": so.loc["01-12-{}".format(y0), "Qkwel"],
                    "J43": bo["hInit_1"],
                    "Rekenblad|G5": bo["EFacMax_1"],
                    "Rekenblad|G6": bo["EFacMin_1"],
                }
                return_dicts.append(onverhard2)
            elif n_onverhard == 2:
                onverhard3 = {
                    "K37": bo["RFacOut_1"],  # RFacOut bakje 3
                    "K38": bo["RFacIn_1"],  # RFacIn bakje 3
                    "K39": bo["por_1"],  # por bakje 3
                    "K40": bo["hMax_1"],  # hMax bakje 3
                    # Qkwel zomer bakje 3
                    "B43": so.loc["01-07-{}".format(y0), "Qkwel"],
                    # Qkwel winter bakje 3
                    "C43": so.loc["01-12-{}".format(y0), "Qkwel"],
                    "K43": bo["hInit_1"],
                    "Rekenblad|H5": bo["EFacMax_1"],
                    "Rekenblad|H6": bo["EFacMin_1"],
                }
                return_dicts.append(onverhard3)
            elif n_onverhard == 3:
                onverhard4 = {
                    "L37": bo["RFacOut_1"],  # RFacOut bakje 4
                    "L38": bo["RFacIn_1"],  # RFacIn bakje 4
                    "L39": bo["por_1"],  # por bakje 4
                    "L40": bo["hMax_1"],  # hMax bakje 4
                    # Qkwel zomer bakje 4
                    "B44": so.loc["01-07-{}".format(y0), "Qkwel"],
                    # Qkwel winter bakje 4
                    "C44": so.loc["01-12-{}".format(y0), "Qkwel"],
                    "L43": bo["hInit_1"],
                    "Rekenblad|I6": bo["EFacMax_1"],
                    "Rekenblad|I5": bo["EFacMin_1"],
                }
                return_dicts.append(onverhard4)
            else:
                print("Warning: only 4 Onverhard buckets can be written to Excel!")
            n_onverhard += 1

    # water bakje
    sw = eag.water.series
    water = {
        "B37": sw.loc["01-07-{}".format(y0), "Qkwel"],
        "B38": sw.loc["01-07-{}".format(y0), "Qwegz"],
        "C37": sw.loc["01-12-{}".format(y0), "Qkwel"],
        "C38": sw.loc["01-12-{}".format(y0), "Qwegz"],
        "B70": eag.water.parameters.at["QInMax_1", "Waarde"],
        "B71": eag.water.parameters.at["QOutMax_1", "Waarde"],
        "D12": eag.water.parameters.at["hBottom_1", "Waarde"],
    }
    return_dicts.append(water)

    # areas
    ms = eag.get_modelstructure()
    areas = {
        "D7": ms.Area.sum(),
        "D8": 0.0,  # verhard
        "D9": 0.0,  # gedraineerd
        "D10": 0.0,  # gemengd gerioleerd
        "D11": eag.water.area,
        "H3": 0.0,  # onverhard 1
        "H4": 0.0,  # onverhard 2
        "H5": 0.0,  # onverhard 3
        "H6": 0.0,  # onverhard 4
    }

    n_onverhard = 0
    for b in buckets:
        if b.startswith("Verhard"):
            areas["D9"] += ms.loc[int(b.split("_")[-1]), "Area"]
        elif b.startswith("Drain"):
            areas["D8"] += ms.loc[int(b.split("_")[-1]), "Area"]
        elif b.startswith("Onverhard"):
            if n_onverhard == 0:
                areas["H3"] += ms.loc[int(b.split("_")[-1]), "Area"]
            elif n_onverhard == 1:
                areas["H4"] += ms.loc[int(b.split("_")[-1]), "Area"]
            elif n_onverhard == 2:
                areas["H5"] += ms.loc[int(b.split("_")[-1]), "Area"]
            elif n_onverhard == 3:
                areas["H6"] += ms.loc[int(b.split("_")[-1]), "Area"]
            n_onverhard += 1
        elif b.startswith("MengRiool"):
            areas["D10"] += ms.loc[int(b.split("_")[-1]), "Area"]
    return_dicts.append(areas)

    return return_dicts

    # # Excel adresses for other parameters
    # chloride = {
    #     "F17": c_neerslag,
    #     "F18": c_kwel,
    #     "F19": c_verhard,
    #     "F20": c_riolering,
    #     "F21": c_gedraineerd,
    #     "F22": c_uitspoeling,
    #     "F23": c_afstroming,
    #     "D33": c_init,
    #     "F24": c_in1,
    #     "F25": c_in2,
    #     "F26": c_in3,
    #     "F27": c_in4,
    #     "F28": c_inPB,
    #     "X75": MP_immissie,
    #     "Z75": MP_1,
    #     "AB75": MP_1,
    #     "AD75": MP_1,
    #     "AF75": MP_1,
    #     "AH75": MP_InlaatPB,
    #     "AJ75": MP_P_Gemaal,
    # }

    # fosfor = {
    #     "D17": p_neerslag,
    #     "D18": p_kwel,
    #     "D19": p_verhard,
    #     "D20": p_riolering,
    #     "D21": p_drain,
    #     "D22": p_uitspoeling,
    #     "D23": p_afstroming,
    #     "D24": p_inlaat1,
    #     "D25": p_inlaat2,
    #     "D26": p_inlaat3,
    #     "D27": p_inlaat4,
    #     "D28": p_inlaatPB,

    #     "E17": p_neerslag_incr,
    #     "E18": p_kwel_incr,
    #     "E19": p_verhard_incr,
    #     "E20": p_riolering_incr,
    #     "E21": p_drain_incr,
    #     "E22": p_uitspoeling_incr,
    #     "E23": p_afstroming_incr,
    #     "E24": p_inlaat1_incr,
    #     "E25": p_inlaat2_incr,
    #     "E26": p_inlaat3_incr,
    #     "E27": p_inlaat4_incr,
    #     "E28": p_inlaatPB_incr,
    # }

    # inuitlaten = {
    #     "A24": naam_inlaat1,
    #     "B24": Qzomer_inlaat1,
    #     "C24": Qwinter_inlaat1,
    #     "A25": naam_inlaat2,
    #     "B25": Qzomer_inlaat2,
    #     "C25": Qwinter_inlaat2,
    #     "A26": naam_inlaat3,
    #     "B26": Qzomer_inlaat3,
    #     "C26": Qwinter_inlaat3,
    #     "A27": naam_inlaat4,
    #     "B27": Qzomer_inlaat4,
    #     "C27": Qwinter_inlaat4,
    #     "A28": naam_inlaat5,
    #     "B28": Qzomer_inlaat5,
    #     "C28": Qwinter_inlaat5,

    #     "A29": naam_uitlaat1,
    #     "B29": Qzomer_uitlaat1,
    #     "C29": Qwinter_uitlaat1,
    #     "A30": naam_uitlaat2,
    #     "B30": Qzomer_uitlaat2,
    #     "C30": Qwinter_uitlaat2,
    #     "A31": naam_uitlaat3,
    #     "B31": Qzomer_uitlaat3,
    #     "C31": Qwinter_uitlaat3,
    #     "A32": naam_uitlaat4,
    #     "B32": Qzomer_uitlaat4,
    #     "C32": Qwinter_uitlaat4,
    #     }


def write_excel(eag, excel_file, write_series=False):

    import openpyxl
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

    newfile = excel_file.split(".")[0] + "_wbpython.xlsx"

    if os.path.isfile(excel_file):
        wb = openpyxl.load_workbook(filename=excel_file)
        ws = wb["uitgangspunten"]
        ws_rekenblad = wb["Rekenblad"]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "uitgangspunten"
        ws_rekenblad = wb.create_sheet("Rekenblad")

    list_of_dicts = eag_params_to_excel_dict(eag)

    # uitgangspunten
    for d in list_of_dicts:
        for k, v in d.items():
            if "|" in k:
                k_split = k.split("|")[-1]
                ws_rekenblad[k_split] = v
            else:
                ws[k] = v

    wb.save(filename=newfile)

    # reeksen
    if write_series:
        series = eag.series

        series_locs = {
            "index": "A77",
            "Neerslag": "B77",
            "Verdamping": "C77",
            "Peil": "D77",
            "Gemaal1Pomp1": "E77",
            "Gemaal1Pomp2": "F77",
            "Gemaal2Pomp1": "G77",
            "Gemaal2Pomp2": "H77",
            "Inlaat voor Calibratie": "I77",
            "q_cso": "J77",
            "Inlaat1": "K77",
            "Inlaat2": "L77",
            "Inlaat3": "M77",
            "Inlaat4": "N77",
            "Uitlaat1": "O77",
            "Uitlaat2": "P77",
            "Uitlaat3": "Q77",
            "Uitlaat4": "R77",
        }

        order = [
            "Neerslag",
            "Verdamping",
            "Peil",
            "Gemaal1Pomp1",
            "Gemaal1Pomp2",
            "Gemaal2Pomp1",
            "Gemaal2Pomp2",
            "Inlaat voor Calibratie",
            "q_cso",
            "Inlaat1",
            "Inlaat2",
            "Inlaat3",
            "Inlaat4",
            "Uitlaat1",
            "Uitlaat2",
            "Uitlaat3",
            "Uitlaat4",
        ]

        # pick only columns that are in eag (in right order)
        has_cols = [col for col in order if col in series.columns]

        # prepare writer for writing series
        xl_writer = pd.ExcelWriter(newfile, engine="openpyxl")
        xl_writer.book = wb
        xl_writer.sheets = {ws.title: ws for ws in wb.worksheets}

        for col in has_cols:
            if col.startswith("Neerslag"):
                series.loc[:, col].to_excel(
                    xl_writer,
                    "uitgangspunten",
                    index=True,
                    header=False,
                    startcol=0,
                    startrow=76,
                )
            else:
                col_letter = coordinate_from_string(series_locs[col])[0]
                startcol = column_index_from_string(col_letter)
                series.loc[:, col].to_excel(
                    xl_writer,
                    "uitgangspunten",
                    index=False,
                    header=False,
                    startcol=startcol - 1,
                    startrow=76,
                )
        xl_writer.save()

    return


def check_numba():
    try:
        from numba import njit as _

        return True
    except ImportError:
        print(
            "Numba is not installed. Installing Numba is "
            "recommended for significant speed-ups."
        )
        return False
    return False


def njit(function):
    try:
        from numba import njit as jit

        return jit(function)
    except ImportError:
        return function
